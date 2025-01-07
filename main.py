from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from office365.runtime.auth.authentication_context import AuthenticationContext
import os
import sys
from dotenv import load_dotenv
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import json

try:
    now = datetime.now()
    formatDateTime = now.strftime("%d-%m-%Y-%H:%M")
except Exception as e:
    with open ('logfile.log', 'a') as file:
        file.write(f"""Problem with date - {str(e)}\n""")
try: 
    local_file = f"test.xlsx"
    # Load env variables
    load_dotenv()
    site_url = os.getenv('site_url')
    file_url = os.getenv('file_url')
    office_username = os.getenv('office_username')
    password = os.getenv('password')

    # db
    db_password = os.getenv('db_password')
    db_user = os.getenv('db_user')
    db_server = os.getenv('db_server')
    db_driver = os.getenv('db_driver')
    db_db = os.getenv('db_db')

    # Email
    from_address = os.getenv('from_address')
    to_address_str = os.getenv('to_address')
    password = os.getenv('password')
except Exception as e:
    with open ('logfile.log', 'a') as file:
        file.write(f"""{formatDateTime} Problem with importing .env data - {str(e)}\n""")  
    sys.exit(0)


#KEYS_TO_DELETE = os.getenv('KEYS_TO_DELETE')
# Download file
def download_file(local_file):
    try:

        # Set download path file
        current_directory = os.path.dirname(os.path.abspath(__file__))
        local_file_path = os.path.join(current_directory, local_file)  # Save file here

        # Authenticate and connect
        credentials = UserCredential(office_username, password)
        ctx = ClientContext(site_url).with_credentials(credentials)
        
        # Download file
        with open(local_file_path, "wb") as file: 
            ctx.web.get_file_by_server_relative_url(file_url).download(file).execute_query()
        print("Plik pobrany.")
        return True
    except Exception as e:
        with open ('logfile.log', 'a') as file:
            file.write(f"""{formatDateTime} Problem with downloading data - {str(e)}\n""")
        return False

# Read downlaoded  file and insert data into db
def read_file(local_file):
    agents_dictionary = {}
    own_dictionary = {}
    coastal_dictionary = {}
    import openpyxl
    with open(local_file, "rb") as file: 
        workbook = openpyxl.load_workbook(file)
        agent_sheets = ['AGENCYJNE','B']
        # AGENTS SHEETS
        print("Agenci:")
        try:
            for sheet_name in agent_sheets:
                sheet = workbook[sheet_name]
                # First half of a file
                counter = 1
                while counter <= sheet.max_row:  # Ensure this iterates over valid rows
                    agents_dictionary[f"{sheet[f'A{counter}'].value}"] = {
                        "city": sheet[f'B{counter}'].value,
                        "phone_number": sheet[f'D{counter}'].value,
                        "address": sheet[f'A{counter + 1}'].value,
                        "salon_email": sheet[f'D{counter + 1}'].value,
                        "agent_data": sheet[f'A{counter + 2}'].value,
                        "agent_email": sheet[f'D{counter + 2}'].value
                    }
                    try:
                        # Remove white spaces from the dictionary key
                        key = f"{sheet[f'A{counter}'].value}".replace(" ", "")
                        agents_dictionary[key] = agents_dictionary.pop(f"{sheet[f'A{counter}'].value}")

                        # Remove spaces from phone_number and agent_email fields
                        agents_dictionary[key]["phone_number"] = agents_dictionary[key]["phone_number"].replace(" ", "")
                        agents_dictionary[key]["agent_email"] = agents_dictionary[key]["agent_email"].replace(" ", "")
                    except Exception as e:
                        pass
                    counter += 3  # Move to the next block of rows for the next agent
                
                # Second half of a file
                counter = 1
                while counter <= sheet.max_row:  # Ensure this iterates over valid rows
                    agents_dictionary[f"{sheet[f'E{counter}'].value}"] = {
                        "city": sheet[f'F{counter}'].value,
                        "phone_number": sheet[f'H{counter}'].value,
                        "address": sheet[f'E{counter + 1}'].value,
                        "salon_email": sheet[f'H{counter + 1}'].value,
                        "agent_data": sheet[f'E{counter + 2}'].value,
                        "agent_email": sheet[f'H{counter + 2}'].value
                    }
                    try:
                        # Remove white spaces from the dictionary key
                        key = f"{sheet[f'A{counter}'].value}".replace(" ", "")
                        agents_dictionary[key] = agents_dictionary.pop(f"{sheet[f'A{counter}'].value}")

                        # Remove spaces from phone_number and agent_email fields
                        agents_dictionary[key]["phone_number"] = agents_dictionary[key]["phone_number"].replace(" ", "")
                        agents_dictionary[key]["agent_email"] = agents_dictionary[key]["agent_email"].replace(" ", "")
                    except Exception as e:
                        pass            
                    counter += 3  # Move to the next block of rows for the next agent
        except Exception as e:
            with open ('logfile.log', 'a') as file:
                file.write(f"""{formatDateTime} Problem with downloading agents data- {str(e)}\n""")   
                
        # OWN SHEET
        sheet = workbook['WŁASNE']
        counter = 4 # FIRST STORE'S OCCURRENCE ROW
        print("Własne:")
        try:
            while counter <= sheet.max_row:  # Ensure this iterates over valid rows
                own_dictionary[f"{sheet[f'A{counter}'].value}"] = {
                    "city": sheet[f'B{counter}'].value,
                    "phone_number": sheet[f'C{counter}'].value,
                    "address": sheet[f'A{counter + 1}'].value,
                    "open_since": sheet[f'A{counter + 2}'].value,
                    "email": sheet[f'C{counter + 2}'].value,
                }
                try:
                    key = f"{sheet[f'A{counter}'].value}".replace(" ", "")
                    own_dictionary[key] = own_dictionary.pop(f"{sheet[f'A{counter}'].value}")
                    own_dictionary[key]["phone_number"] = own_dictionary[key]["phone_number"].replace(" ", "")
                    own_dictionary[key]["email"] = own_dictionary[key]["email"].replace(" ", "")
                except:
                    #None would be ignored
                    pass
        
                counter += 3  # Move to the next block of rows for the next agent
            # Second half of a file
            counter = 4 # FIRST STORE'S OCCURRENCE ROW
            while counter <= sheet.max_row:  # Ensure this iterates over valid rows
                own_dictionary[f"{sheet[f'D{counter}'].value}"] = {
                    "city": sheet[f'E{counter}'].value,
                    "phone_number": sheet[f'F{counter}'].value,
                    "address": sheet[f'D{counter + 1}'].value,
                    "open_since": sheet[f'D{counter + 2}'].value,
                    "email": sheet[f'F{counter + 2}'].value,
                }
                
                try:
                    key = f"{sheet[f'A{counter}'].value}".replace(" ", "")
                    own_dictionary[key] = own_dictionary.pop(f"{sheet[f'A{counter}'].value}")
                    own_dictionary[key]["phone_number"] = own_dictionary[key]["phone_number"].replace(" ", "")
                    own_dictionary[key]["email"] = own_dictionary[key]["email"].replace(" ", "")
                except:
                    #None would be ignored
                    pass
        
                counter += 3  # Move to the next block of rows for the next agent
        except Exception as e:
            with open ('logfile.log', 'a') as file:
                file.write(f"""{formatDateTime} Problem with downloading own data- {str(e)}\n""")  
                
        # COASTAL
        sheet = workbook['NADMORSKIE']
        counter = 1 # FIRST STORE'S OCCURRENCE ROW
        print("Nadmorskie:")
        try:
            while counter <= sheet.max_row:  # Ensure this iterates over valid rows
                coastal_dictionary[f"{sheet[f'A{counter}'].value}"] = {
                    "agent_number": sheet[f'C{counter}'].value,
                    "address": sheet[f'A{counter + 1}'].value,
                    "email": sheet[f'D{counter + 1}'].value,
                    "agent_data": sheet[f'A{counter + 2}'].value,
                    "agent_email": sheet[f'D{counter + 2}'].value,
                }
                
                try:
                    # Remove white spaces from the dictionary key
                    key = f"{sheet[f'A{counter}'].value}".replace(" ", "")
                    coastal_dictionary[key] = coastal_dictionary.pop(f"{sheet[f'A{counter}'].value}")

                    # Remove spaces from email fields
                    coastal_dictionary[key]["email"] = coastal_dictionary[key]["email"].replace(" ", "")
                    coastal_dictionary[key]["agent_email"] = coastal_dictionary[key]["agent_email"].replace(" ", "")
                except Exception as e:
                    print(e)
                    print("-----")
                    print(counter)
        
                counter += 3  # Move to the next block of rows for the next agent
        except Exception as e:
            with open ('logfile.log', 'a') as file:
                file.write(f"""{formatDateTime} Problem with downloading coastal data - {str(e)}\n""")  
            
    # Delete None key if exists in agents
    try:
        del agents_dictionary['None']
    except:
        pass

    try:
        del coastal_dictionary['None']
    except:
        pass
    try:
        del own_dictionary['None']
    except:
        pass
    
    # Delete keys longer than 4 chars (shop code)
    print("Usuwanie kluczy dłuższych niz 4:")
    for key in list(own_dictionary.keys()):  # list() is used to avoid modifying the dictionary while iterating
        if len(key) > 4:
            del own_dictionary[key]
    

    
    # Insert into db
    print("Wgyrwanie danych do db:")
    try:
        import pyodbc
        conn = pyodbc.connect(f'DRIVER={db_driver};'
                      f'SERVER={db_server};'
                      f'DATABASE={db_db};'
                      f'UID={db_user};'
                      f'PWD={db_password}')
        cursor = conn.cursor()
        cursor.execute("TRUNCATE TABLE contact_details_own")
        for key, value in own_dictionary.items():
            cursor.execute("""
                INSERT INTO contact_details_own (code, city, phone_number, address, open_since, email)
                VALUES (?, ?, ?, ?, ?, ?)
                """, key, value['city'], value['phone_number'], value['address'], value['open_since'], value['email'])

        conn.commit()
        
        cursor.execute("TRUNCATE TABLE contact_details_agent")
        for key, value in agents_dictionary.items():
            cursor.execute("""
                INSERT INTO contact_details_agent (code, city, phone_number, address, salon_email, agent_data, agent_email)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, key, value['city'], value['phone_number'], value['address'], value['salon_email'], value['agent_data'], value['agent_email'])
        conn.commit()
        
        cursor.execute("TRUNCATE TABLE contact_details_coastal")
        for key, value in coastal_dictionary.items():
            cursor.execute("""
                INSERT INTO contact_details_coastal (code, agent_number, address, email, agent_data, agent_email)
                VALUES (?, ?, ?, ?, ?, ?)
            """, key, value['agent_number'], value['address'], value['email'], value['agent_data'], value['agent_email'])

        # Committing the transaction
        conn.commit()
        print("Done.")
        send_mail(True,len(own_dictionary),len(agents_dictionary),len(coastal_dictionary))
    except Exception as e:
        with open ('logfile.log', 'a') as file:
            file.write(f"""{formatDateTime} Problem with inserting data to db- {str(e)}\n""")
        send_mail(False,0,0,0)
    cursor.close()
    conn.close()
        
def send_mail(status,own_count,agent_count,coastal_count):
    print("Wysyłka Email:")
    to_address = json.loads(to_address_str)
    msg = MIMEMultipart()
    msg['From'] = from_address
    msg["To"] = ", ".join(to_address)
    msg['Subject'] = f"Zgranie danych teleadresowych do bazy: {formatDateTime}."
    #print(", ".join(to_address))
    #body = ""
    if status:
        total_count = own_count + agent_count + coastal_count
        body = f"Pobrane dane salony własne: {own_count}, salony agencyjne: {agent_count}, salony nadmorskie: {coastal_count}\nŁacznie: {total_count}"
        msg.attach(MIMEText(body, 'html'))
    else:
        body = f"Problem z działaniem skryptu. Sprawdź logi."
        msg.attach(MIMEText(body, 'html'))
    try:
        server = smtplib.SMTP('smtp-mail.outlook.com', 587)
        server.starttls()
        server.login(from_address, password)
        text = msg.as_string()
        server.sendmail(from_address, to_address, text)
        server.quit()               
    except Exception as e:
        with open ('logfile.log', 'a') as file:
            file.write(f"""{formatDateTime} Problem z wysłaniem na maile\n{str(e)}\n""")

    
# Run script only if file is downloaded
if download_file(local_file):
    read_file(local_file)



#read_file(local_file)




